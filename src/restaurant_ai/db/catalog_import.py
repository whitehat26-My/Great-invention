"""Load a real menu from a spreadsheet.

Everything the agents report is computed from the catalog: recipes explode into
ingredient demand, ingredient costs roll up into plate costs, plate costs into
margins, margins into Irma's classifications and Camelia's verdicts. Until the
catalog is the restaurant's own, every number the platform produces describes a
fiction — and the only way in was SQL.

This is the way in. ``write_template`` produces a workbook whose example rows
are themselves a valid import (documentation that cannot drift from the code,
because the test suite imports it), and ``import_catalog`` loads a filled-in
copy: suppliers, ingredients with their preferred pack, sub-recipes, bills of
materials, menu items.

Two properties matter more than the parsing:

- **All-or-nothing.** Every error in the file is reported at once, with its
  sheet and row, and nothing is written until the whole file is clean. A
  spreadsheet is fixed in batches; an importer that stops at the first mistake
  and half-loads the rest turns one bad row into a corrupt catalog.
- **It ends with proof.** After loading, every menu item is costed through the
  same recipe explosion the agents use. A dish that cannot be costed aborts the
  import — better to hear it now than to have Betrisha discover it at six in
  the morning.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from restaurant_ai.db.base import money, qty
from restaurant_ai.db.models import (
    Allergen,
    Availability,
    Ingredient,
    MenuItem,
    MenuSection,
    Recipe,
    RecipeComponent,
    ShiftRole,
    Staff,
    Station,
    StockItem,
    Supplier,
)
from restaurant_ai.db.seed import _upsert
from restaurant_ai.logging_setup import get_logger

log = get_logger(__name__)

SHEETS = ("Suppliers", "Staff", "Ingredients", "SubRecipes", "BOM", "Menu", "Allergens")

_STATIONS = {station.value for station in Station}
_ROLES = {role.value for role in ShiftRole}


def _clock(text: str) -> time:
    """ "10:00" as a time. Spreadsheets hand this back three different ways."""
    if isinstance(text, time):
        return text
    cleaned = str(text).strip()
    hour, _, minute = cleaned.partition(":")
    return time(int(hour), int(minute or 0))


class CatalogImportError(Exception):
    """The file has problems. All of them are listed, none of it was written."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__(
            f"{len(errors)} problem(s) in the workbook; nothing was imported:\n"
            + "\n".join(f"  - {e}" for e in errors)
        )


@dataclass
class ImportSummary:
    counts: dict[str, int] = field(default_factory=dict)
    costings: list[dict[str, Any]] = field(default_factory=list)
    deactivated: list[str] = field(default_factory=list)
    uncosted: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------


def _rows(workbook, sheet: str) -> list[tuple[int, dict[str, Any]]]:
    """Rows of one sheet as dicts keyed by header, with worksheet row numbers.

    The row number is kept because it is the only address the person fixing the
    file has — "Ingredients row 14" is actionable, "the third bad row" is not.
    """
    if sheet not in workbook.sheetnames:
        return []
    ws = workbook[sheet]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in header_cells]

    out: list[tuple[int, dict[str, Any]]] = []
    for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {h: v for h, v in zip(headers, values, strict=False) if h}
        if all(v is None or str(v).strip() == "" for v in row.values()):
            continue  # a blank line in a spreadsheet is formatting, not data
        out.append((index, row))
    return out


def _text(row: dict[str, Any], key: str) -> str:
    value = row.get(key)
    return "" if value is None else str(value).strip()


def _number(row: dict[str, Any], key: str, errors: list[str], where: str) -> Decimal | None:
    raw = row.get(key)
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return Decimal(str(raw))
    except InvalidOperation:
        errors.append(f"{where}: '{key}' is not a number ({raw!r}).")
        return None


def _integer(row: dict[str, Any], key: str, errors: list[str], where: str) -> int | None:
    value = _number(row, key, errors, where)
    return None if value is None else int(value)


# ---------------------------------------------------------------------------
# Importing
# ---------------------------------------------------------------------------


def import_catalog(
    session: Session,
    path: str | Path,
    replace_menu: bool = False,
    allow_uncosted: bool = False,
) -> ImportSummary:
    """Validate the whole workbook, then load it, then prove it costs out.

    Raises :class:`CatalogImportError` with every problem found; on any raise
    the session has had no lasting writes flushed past the caller's transaction.

    ``allow_uncosted`` admits dishes with no bill of materials. A restaurant
    knows its own prices long before it has costed a single recipe, and refusing
    the menu until every ingredient is priced keeps the real prices out and
    leaves the demo ones in — which is worse than an incomplete catalog. The
    dishes are loaded, listed in ``summary.uncosted``, and excluded from margin
    analysis rather than counted as costing nothing; off by default, because a
    catalog that claims to be complete should be.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    errors: list[str] = []
    summary = ImportSummary()

    # --- Allergens: the vocabulary everything else's allergen column uses ---
    known_allergens = set(session.execute(select(Allergen.code)).scalars())
    for line, row in _rows(workbook, "Allergens"):
        where = f"Allergens row {line}"
        code, label = _text(row, "code").lower(), _text(row, "label")
        if not code:
            errors.append(f"{where}: 'code' is required.")
            continue
        _upsert(session, Allergen, {"code": code}, {"label": label or code.title()})
        known_allergens.add(code)

    # --- Suppliers -----------------------------------------------------------
    suppliers: dict[str, Supplier] = {
        s.code: s for s in session.execute(select(Supplier)).scalars()
    }
    for line, row in _rows(workbook, "Suppliers"):
        where = f"Suppliers row {line}"
        code, name = _text(row, "code"), _text(row, "name")
        if not code or not name:
            errors.append(f"{where}: 'code' and 'name' are required.")
            continue
        suppliers[code] = _upsert(
            session,
            Supplier,
            {"code": code},
            {
                "name": name,
                "email": _text(row, "email") or None,
                "phone": _text(row, "phone") or None,
                "lead_time_days": _integer(row, "lead_time_days", errors, where) or 2,
                "min_order_value": money(_number(row, "min_order_value", errors, where) or 0),
                "delivery_days": _text(row, "delivery_days") or "0,1,2,3,4,5,6",
                "is_active": True,
            },
        )
    session.flush()

    # --- Staff, and when they can work ---------------------------------------
    # Henry could never be given a roster to build. `readiness` told the owner
    # "Henry needs who works here, their roles, and when they can work" and
    # there was no sheet, no command and no way to say it — advice for a thing
    # that could not be done.
    #
    # Availability lives on the same row rather than a sheet of its own. A
    # mamak's staff work the same hours most days, and a second sheet keyed by
    # employee code is a second chance to typo the code.
    for line, row in _rows(workbook, "Staff"):
        where = f"Staff row {line}"
        code, name = _text(row, "code"), _text(row, "name")
        if not code or not name:
            errors.append(f"{where}: 'code' and 'name' are required.")
            continue
        role = _text(row, "role").lower().replace(" ", "_")
        if role not in _ROLES:
            errors.append(f"{where}: role {role!r} is not one of {', '.join(sorted(_ROLES))}.")
            continue

        member = _upsert(
            session,
            Staff,
            {"employee_code": code},
            {
                "name": name,
                "role": role,
                "hourly_rate": money(_number(row, "hourly_rate", errors, where) or 0),
                "max_weekly_hours": _integer(row, "max_weekly_hours", errors, where) or 48,
                "min_rest_hours": _integer(row, "min_rest_hours", errors, where) or 11,
                "phone": _text(row, "phone") or None,
                "is_active": True,
            },
        )
        session.flush()

        days = _text(row, "days") or "1,2,3,4,5,6"
        start, end = _text(row, "from") or "10:00", _text(row, "to") or "23:00"
        try:
            opens, closes = _clock(start), _clock(end)
        except ValueError:
            errors.append(f"{where}: 'from'/'to' must look like 10:00, not {start!r}/{end!r}.")
            continue
        # Rewritten wholesale, so an edited spreadsheet is the truth rather than
        # something added to what was there before.
        #
        # Deleted and flushed rather than cleared: `clear()` leaves the removals
        # pending, and SQLAlchemy inserts the new rows before it issues the
        # deletes — so re-importing an unchanged sheet collides with the unique
        # constraint on (staff, weekday, start).
        session.execute(delete(Availability).where(Availability.staff_id == member.id))
        session.flush()
        for part in days.split(","):
            part = part.strip()
            if not part:
                continue
            if not part.isdigit() or not 0 <= int(part) <= 6:
                errors.append(f"{where}: 'days' takes 0-6 (0 = Monday), not {part!r}.")
                continue
            member.availability.append(
                Availability(weekday=int(part), start_time=opens, end_time=closes)
            )
    session.flush()

    # --- Ingredients (with their preferred supplier pack) --------------------
    ingredients: dict[str, Ingredient] = {
        i.code: i for i in session.execute(select(Ingredient)).scalars()
    }
    pack_rows: list[tuple[str, str, dict[str, Any]]] = []
    for line, row in _rows(workbook, "Ingredients"):
        where = f"Ingredients row {line}"
        code, name, unit = _text(row, "code"), _text(row, "name"), _text(row, "unit").lower()
        if not code or not name or not unit:
            errors.append(f"{where}: 'code', 'name' and 'unit' are required.")
            continue
        if unit not in ("g", "ml", "ea"):
            errors.append(f"{where}: 'unit' must be g, ml or ea, not {unit!r}.")
            continue

        yield_pct = _number(row, "yield_pct", errors, where)
        if yield_pct is not None and not (0 < yield_pct <= 1):
            errors.append(
                f"{where}: 'yield_pct' must be a fraction — 0.95 means 5% is lost "
                f"to trim, not {yield_pct}."
            )
            continue

        allergens = [a.strip().lower() for a in _text(row, "allergens").split(",") if a.strip()]
        unknown = [a for a in allergens if a not in known_allergens]
        if unknown:
            errors.append(
                f"{where}: unknown allergen(s) {', '.join(unknown)}. Declare new "
                f"ones on the Allergens sheet; known: {', '.join(sorted(known_allergens))}."
            )
            continue

        ingredients[code] = _upsert(
            session,
            Ingredient,
            {"code": code},
            {
                "name": name,
                "base_uom": unit,
                "cost_per_base_unit": qty(_number(row, "cost_per_unit", errors, where) or 0),
                "yield_pct": qty(yield_pct if yield_pct is not None else 1),
                "shelf_life_days": _integer(row, "shelf_life_days", errors, where) or 7,
                "allergen_codes": ",".join(allergens),
            },
        )

        supplier_code = _text(row, "supplier_code")
        if supplier_code:
            if supplier_code not in suppliers:
                errors.append(
                    f"{where}: supplier_code {supplier_code!r} is not on the Suppliers sheet."
                )
                continue
            pack_size = _number(row, "pack_size", errors, where)
            pack_price = _number(row, "pack_price", errors, where)
            if not pack_size or pack_price is None:
                errors.append(
                    f"{where}: a supplier_code needs 'pack_size' (in {unit}) and 'pack_price' "
                    f"— they are what a purchase order is written in."
                )
                continue
            pack_rows.append(
                (
                    code,
                    supplier_code,
                    {
                        "supplier_sku": _text(row, "supplier_sku") or code,
                        "pack_size": qty(pack_size),
                        "pack_uom": unit,
                        "contract_price": money(pack_price),
                        "min_order_qty": qty(_number(row, "min_order_packs", errors, where) or 1),
                        "is_preferred": True,
                    },
                )
            )
    session.flush()

    for ing_code, sup_code, values in pack_rows:
        supplier_sku = values.pop("supplier_sku")
        _upsert(
            session,
            StockItem,
            {"supplier_id": suppliers[sup_code].id, "supplier_sku": supplier_sku},
            {"ingredient_id": ingredients[ing_code].id, **values},
        )
    session.flush()

    # --- Sub-recipes (shells first; their BOMs arrive with everyone else's) --
    recipes: dict[str, Recipe] = {r.code: r for r in session.execute(select(Recipe)).scalars()}
    sub_recipe_codes: set[str] = set()
    for line, row in _rows(workbook, "SubRecipes"):
        where = f"SubRecipes row {line}"
        code, name = _text(row, "code"), _text(row, "name")
        if not code or not name:
            errors.append(f"{where}: 'code' and 'name' are required.")
            continue
        station = _text(row, "station").lower() or Station.SAUTE.value
        if station not in _STATIONS:
            errors.append(f"{where}: station must be one of {', '.join(sorted(_STATIONS))}.")
            continue
        yield_qty = _number(row, "yield_qty", errors, where)
        yield_uom = _text(row, "yield_uom").lower()
        if not yield_qty or yield_uom not in ("g", "ml", "ea"):
            errors.append(
                f"{where}: 'yield_qty' and 'yield_uom' (g/ml/ea) are required — a batch "
                f"has to say what it makes for plated recipes to draw from it."
            )
            continue
        recipes[code] = _upsert(
            session,
            Recipe,
            {"code": code},
            {
                "name": name,
                "yield_qty": qty(yield_qty),
                "yield_uom": yield_uom,
                "station": Station(station),
                "prep_seconds": (_integer(row, "prep_minutes", errors, where) or 5) * 60,
                "method": _text(row, "method") or None,
                "menu_item_id": None,
            },
        )
        sub_recipe_codes.add(code)
    session.flush()

    # --- Menu items, each with its own plated recipe -------------------------
    sections: dict[str, MenuSection] = {
        s.name: s for s in session.execute(select(MenuSection)).scalars()
    }
    menu_skus: set[str] = set()
    menu_recipe_by_sku: dict[str, Recipe] = {}
    for line, row in _rows(workbook, "Menu"):
        where = f"Menu row {line}"
        sku, name, section_name = _text(row, "sku"), _text(row, "name"), _text(row, "section")
        if not sku or not name or not section_name:
            errors.append(f"{where}: 'sku', 'name' and 'section' are required.")
            continue
        if sku in menu_skus:
            errors.append(f"{where}: sku {sku!r} appears twice in the Menu sheet.")
            continue
        price = _number(row, "price", errors, where)
        if price is None or price <= 0:
            errors.append(f"{where}: 'price' must be a positive number.")
            continue
        station = _text(row, "station").lower() or Station.SAUTE.value
        if station not in _STATIONS:
            errors.append(f"{where}: station must be one of {', '.join(sorted(_STATIONS))}.")
            continue

        if section_name not in sections:
            sections[section_name] = _upsert(
                session, MenuSection, {"name": section_name}, {"display_order": len(sections)}
            )
            session.flush()

        item = _upsert(
            session,
            MenuItem,
            {"sku": sku},
            {
                "name": name,
                "description": _text(row, "description") or None,
                "section_id": sections[section_name].id,
                "price": money(price),
                "station": Station(station),
                "prep_seconds": (_integer(row, "prep_minutes", errors, where) or 8) * 60,
                "course": _integer(row, "course", errors, where) or 2,
                "is_active": True,
            },
        )
        session.flush()
        # The same convention the seed uses, so imported and seeded dishes
        # cost out through identical plumbing.
        recipe_code = f"REC-{sku.split('-', 1)[1]}" if "-" in sku else f"REC-{sku}"
        recipe = _upsert(
            session,
            Recipe,
            {"code": recipe_code},
            {
                "name": name,
                "menu_item_id": item.id,
                "yield_qty": qty("1"),
                "yield_uom": "ea",
                "station": Station(station),
                "prep_seconds": item.prep_seconds,
            },
        )
        menu_skus.add(sku)
        menu_recipe_by_sku[sku] = recipe
        recipes[recipe_code] = recipe
    session.flush()

    # --- Bills of materials ---------------------------------------------------
    # Parents named by sub-recipe code or menu SKU; components by ingredient or
    # sub-recipe code. Grouped so each parent's components are replaced whole:
    # a re-import must never append a second copy of a recipe.
    bom: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for line, row in _rows(workbook, "BOM"):
        parent = _text(row, "parent_code")
        if not parent:
            errors.append(f"BOM row {line}: 'parent_code' is required.")
            continue
        bom.setdefault(parent, []).append((line, row))

    for parent, entries in bom.items():
        if parent in menu_recipe_by_sku:
            recipe = menu_recipe_by_sku[parent]
        elif parent in sub_recipe_codes:
            recipe = recipes[parent]
        else:
            lines = ", ".join(str(line) for line, _ in entries)
            errors.append(
                f"BOM rows {lines}: parent_code {parent!r} is neither a Menu sku nor a "
                f"SubRecipes code in this workbook."
            )
            continue

        components: list[RecipeComponent] = []
        for line, row in entries:
            where = f"BOM row {line}"
            component_code = _text(row, "component_code")
            quantity = _number(row, "quantity", errors, where)
            uom = _text(row, "uom").lower()
            if not component_code or not quantity or quantity <= 0:
                errors.append(f"{where}: 'component_code' and a positive 'quantity' are required.")
                continue

            if component_code in sub_recipe_codes:
                target_uom = recipes[component_code].yield_uom
                if uom != target_uom:
                    errors.append(
                        f"{where}: {component_code} yields {target_uom!r} but is used as "
                        f"{uom!r}; unit conversion is not supported here, use {target_uom!r}."
                    )
                    continue
                components.append(
                    RecipeComponent(
                        sub_recipe_id=recipes[component_code].id,
                        quantity=qty(quantity),
                        uom=uom,
                        note=_text(row, "note") or None,
                    )
                )
            elif component_code in ingredients:
                target_uom = ingredients[component_code].base_uom
                if uom != target_uom:
                    errors.append(
                        f"{where}: {component_code} is measured in {target_uom!r} but is used "
                        f"as {uom!r}; unit conversion is not supported here, use {target_uom!r}."
                    )
                    continue
                components.append(
                    RecipeComponent(
                        ingredient_id=ingredients[component_code].id,
                        quantity=qty(quantity),
                        uom=uom,
                        note=_text(row, "note") or None,
                    )
                )
            else:
                errors.append(
                    f"{where}: component_code {component_code!r} is not an Ingredients code "
                    f"or a SubRecipes code."
                )

        if not errors:
            recipe.components.clear()
            session.flush()
            recipe.components.extend(components)
    session.flush()

    # A menu item with no BOM cannot be costed, forecast or deducted from
    # stock. It is a picture of a dish, not a dish.
    for sku, recipe in menu_recipe_by_sku.items():
        if sku not in bom and not recipe.components:
            if allow_uncosted:
                summary.uncosted.append(sku)
                continue
            errors.append(
                f"Menu sku {sku!r} has no BOM rows — without one it cannot be costed, "
                f"forecast, or deducted from stock. Pass allow_uncosted to load it anyway."
            )

    if errors:
        raise CatalogImportError(errors)

    # --- Optionally retire everything the file no longer mentions ------------
    if replace_menu and menu_skus:
        for item in session.execute(select(MenuItem).where(MenuItem.is_active)).scalars():
            if item.sku not in menu_skus:
                item.is_active = False
                summary.deactivated.append(item.sku)
    session.flush()

    # --- The proof: every imported dish must cost out ------------------------
    from restaurant_ai.domain.costing import cost_breakdown, menu_item_allergens

    costing_errors: list[str] = []
    for sku in sorted(menu_skus - set(summary.uncosted)):
        item = session.execute(select(MenuItem).where(MenuItem.sku == sku)).scalar_one()
        try:
            breakdown = cost_breakdown(session, item.id)
            dish_allergens = menu_item_allergens(session, item.id)
        except Exception as exc:
            costing_errors.append(f"Menu sku {sku!r} does not cost out: {exc}")
            continue
        plate_cost = breakdown.total_cost
        margin_pct = (item.price - plate_cost) / item.price if item.price else Decimal("0")
        summary.costings.append(
            {
                "sku": sku,
                "name": item.name,
                "price": item.price,
                "plate_cost": plate_cost.quantize(Decimal("0.01")),
                "margin_pct": margin_pct.quantize(Decimal("0.001")),
                "allergens": sorted(dish_allergens),
            }
        )
    if costing_errors:
        raise CatalogImportError(costing_errors)

    summary.counts = {
        "suppliers": len(_rows(workbook, "Suppliers")),
        "staff": len(_rows(workbook, "Staff")),
        "ingredients": len(_rows(workbook, "Ingredients")),
        "sub_recipes": len(sub_recipe_codes),
        "menu_items": len(menu_skus),
        "bom_lines": sum(len(v) for v in bom.values()),
    }
    log.info("catalog imported", **summary.counts)
    return summary


# ---------------------------------------------------------------------------
# The template
# ---------------------------------------------------------------------------

_README = [
    ("How to fill this in", ""),
    ("", ""),
    (
        "Suppliers",
        "Who you buy from. lead_time_days drives reorder points; delivery_days is 0=Mon..6=Sun.",
    ),
    (
        "Ingredients",
        "One row per raw good, in its base unit (g, ml or ea). cost_per_unit is per ONE base unit — RM per gram, not per kilo.",
    ),
    (
        "",
        "yield_pct is the usable fraction: 0.85 means 15% is lost to trim. allergens is a comma list of codes from the Allergens sheet.",
    ),
    (
        "",
        "supplier_code + pack_size + pack_price describe how you actually buy it: a 25000 g sack for RM 62.50 is pack_size 25000, pack_price 62.50.",
    ),
    (
        "SubRecipes",
        "Batches made ahead — a sambal, a stock, a spice paste. yield_qty/yield_uom is what one batch makes; plated recipes draw from it.",
    ),
    (
        "BOM",
        "What goes into what. parent_code is a Menu sku or a SubRecipes code; component_code is an Ingredients or SubRecipes code.",
    ),
    (
        "",
        "quantity is in the component's own unit. Every Menu sku needs at least one BOM row — a dish with no recipe cannot be costed.",
    ),
    (
        "Menu",
        "What guests order. price is what they pay; station is where it is cooked (grill, fry, saute, cold, pastry, bar); course 1=starter 2=main 3=dessert.",
    ),
    ("Allergens", "Only needed to add codes beyond the built-in list."),
    ("", ""),
    (
        "The example rows in every sheet are a working import — try `restaurant-ai import-menu` on this file untouched.",
        "",
    ),
    (
        "Importing is all-or-nothing: every problem is reported with its sheet and row, and nothing loads until the file is clean.",
        "",
    ),
]

_TEMPLATE: dict[str, tuple[list[str], list[list[Any]]]] = {
    "Suppliers": (
        ["code", "name", "email", "phone", "lead_time_days", "min_order_value", "delivery_days"],
        [
            [
                "SUP-EXAMPLE",
                "Pasar Segar Sdn Bhd",
                "orders@pasarsegar.example",
                "+60 3-0000 0000",
                2,
                150,
                "0,2,4",
            ],
        ],
    ),
    "Staff": (
        [
            "code",
            "name",
            "role",
            "hourly_rate",
            "max_weekly_hours",
            "min_rest_hours",
            "phone",
            "days",
            "from",
            "to",
        ],
        [
            # Availability on the same row: a mamak's staff work the same hours
            # most days, and a second sheet keyed by employee code is a second
            # chance to mistype the code.
            [
                "EMP-001",
                "Ahmad",
                "server",
                9.50,
                48,
                11,
                "+60 12-000 0000",
                "0,1,2,3,4,5",
                "10:00",
                "23:00",
            ],
            ["EMP-002", "Siti", "chef", 14.00, 48, 11, "", "1,2,3,4,5,6", "09:00", "22:00"],
        ],
    ),
    "Ingredients": (
        [
            "code",
            "name",
            "unit",
            "cost_per_unit",
            "yield_pct",
            "shelf_life_days",
            "allergens",
            "supplier_code",
            "pack_size",
            "pack_price",
            "supplier_sku",
            "min_order_packs",
        ],
        [
            [
                "ING-RICE",
                "Jasmine rice",
                "g",
                0.0045,
                1.0,
                180,
                "",
                "SUP-EXAMPLE",
                10000,
                45,
                "RICE-10KG",
                1,
            ],
            [
                "ING-CHICKEN",
                "Chicken thigh, boneless",
                "g",
                0.016,
                0.92,
                3,
                "",
                "SUP-EXAMPLE",
                2000,
                32,
                "CHK-2KG",
                1,
            ],
            [
                "ING-PEANUT",
                "Peanuts, roasted",
                "g",
                0.012,
                1.0,
                60,
                "peanut",
                "SUP-EXAMPLE",
                1000,
                12,
                "PNT-1KG",
                1,
            ],
            [
                "ING-CHILI",
                "Dried chili",
                "g",
                0.02,
                1.0,
                120,
                "",
                "SUP-EXAMPLE",
                500,
                10,
                "CHL-500",
                1,
            ],
        ],
    ),
    "SubRecipes": (
        ["code", "name", "yield_qty", "yield_uom", "station", "prep_minutes", "method"],
        [
            [
                "SUB-SAMBAL",
                "House sambal",
                1000,
                "g",
                "saute",
                45,
                "Blend chilies, fry low until the oil splits.",
            ],
        ],
    ),
    "BOM": (
        ["parent_code", "component_code", "quantity", "uom", "note"],
        [
            ["SUB-SAMBAL", "ING-CHILI", 300, "g", ""],
            ["SUB-SAMBAL", "ING-PEANUT", 100, "g", "ground in"],
            ["MNU-EXAMPLE", "ING-RICE", 180, "g", "steamed"],
            ["MNU-EXAMPLE", "ING-CHICKEN", 150, "g", ""],
            ["MNU-EXAMPLE", "SUB-SAMBAL", 40, "g", "on the side"],
        ],
    ),
    "Menu": (
        ["sku", "name", "section", "price", "description", "station", "course", "prep_minutes"],
        [
            [
                "MNU-EXAMPLE",
                "Example Chicken Rice",
                "Mains",
                15.90,
                "Replace me with your first real dish.",
                "grill",
                2,
                9,
            ],
        ],
    ),
    "Allergens": (
        ["code", "label"],
        [
            ["peanut", "Peanut"],
        ],
    ),
}


def write_template(path: str | Path) -> Path:
    """Write the fill-in workbook.

    The example rows are a coherent, importable catalog on purpose: they are
    the documentation, and because the test suite imports this exact file, the
    documentation cannot quietly drift from what the importer accepts.
    """
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    readme = workbook.active
    readme.title = "ReadMe"
    for title, note in _README:
        readme.append([title, note])
    readme.column_dimensions["A"].width = 40
    readme.column_dimensions["B"].width = 120
    for row in readme.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)

    for sheet, (headers, examples) in _TEMPLATE.items():
        ws = workbook.create_sheet(sheet)
        ws.append(headers)
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            cell.font = Font(bold=True)
        for example in examples:
            ws.append(example)
        for column, header in zip("ABCDEFGHIJKL", headers, strict=False):
            ws.column_dimensions[column].width = max(14, len(header) + 4)

    out = Path(path)
    workbook.save(out)
    return out
