"""The spreadsheet importer — the way a real menu gets in without SQL.

The template's example rows are the documentation, so the first test imports
the template untouched: if the docs and the importer ever disagree, this file
is where it shows up.
"""

from __future__ import annotations

from decimal import Decimal

import pytest
from sqlalchemy import select
from typer.testing import CliRunner

from restaurant_ai.db.catalog_import import (
    CatalogImportError,
    import_catalog,
    write_template,
)
from restaurant_ai.db.models import (
    Ingredient,
    MenuItem,
    Recipe,
    StockItem,
    Supplier,
)

pytestmark = pytest.mark.db


@pytest.fixture
def template(tmp_path):
    return write_template(tmp_path / "menu.xlsx")


def _workbook(template, mutate):
    """Load the template, let the test break it, save the broken copy."""
    import openpyxl

    wb = openpyxl.load_workbook(template)
    mutate(wb)
    out = template.parent / "mutated.xlsx"
    wb.save(out)
    return out


class TestTheTemplateIsTrue:
    """The example rows must be a working import, forever."""

    def test_the_untouched_template_imports(self, db, template):
        summary = import_catalog(db, template)

        assert summary.counts == {
            "suppliers": 1,
            "staff": 2,
            "ingredients": 4,
            "sub_recipes": 1,
            "menu_items": 1,
            "bom_lines": 5,
        }
        item = db.execute(select(MenuItem).where(MenuItem.sku == "MNU-EXAMPLE")).scalar_one()
        assert item.price == Decimal("15.90")

    def test_it_ends_with_a_costed_dish(self, db, template):
        summary = import_catalog(db, template)
        (costing,) = summary.costings
        assert costing["sku"] == "MNU-EXAMPLE"
        assert Decimal("0") < costing["plate_cost"] < costing["price"]

    def test_allergens_climb_through_the_sub_recipe(self, db, template):
        """The peanut is in the sambal, not the dish — it must surface anyway.

        This is the recursive BOM walk doing the one job that matters most:
        an allergen two levels down is still the dish's allergen.
        """
        summary = import_catalog(db, template)
        assert "peanut" in summary.costings[0]["allergens"]

    def test_the_supplier_pack_arrives(self, db, template):
        import_catalog(db, template)
        ingredient = db.execute(
            select(Ingredient).where(Ingredient.code == "ING-RICE")
        ).scalar_one()
        pack = db.execute(
            select(StockItem).where(StockItem.ingredient_id == ingredient.id)
        ).scalar_one()
        assert pack.pack_size == Decimal("10000")
        assert pack.contract_price == Decimal("45.00")

    def test_importing_twice_changes_nothing(self, db, template):
        """Re-import must update in place, never duplicate.

        The failure mode is the BOM: components are child rows, and appending
        a second copy silently doubles every plate cost and prep quantity.
        """
        import_catalog(db, template)
        first = db.execute(select(Recipe).where(Recipe.code == "REC-EXAMPLE")).scalar_one()
        count_before = len(first.components)
        cost_before = import_catalog(db, template).costings[0]["plate_cost"]

        db.expire_all()
        again = db.execute(select(Recipe).where(Recipe.code == "REC-EXAMPLE")).scalar_one()
        assert len(again.components) == count_before
        assert cost_before == import_catalog(db, template).costings[0]["plate_cost"]


class TestNothingLoadsUntilEverythingIsClean:
    def test_every_error_is_reported_at_once(self, db, template):
        """A spreadsheet is fixed in batches, so all problems come back together."""

        def break_many(wb):
            wb["Ingredients"]["H2"] = "SUP-NOWHERE"  # unknown supplier
            wb["Ingredients"]["E3"] = 95  # yield_pct as a percent, not a fraction
            wb["Ingredients"]["G4"] = "penut"  # misspelled allergen
            wb["Menu"]["D2"] = -5  # negative price

        broken = _workbook(template, break_many)
        with pytest.raises(CatalogImportError) as caught:
            import_catalog(db, broken)

        text = str(caught.value)
        assert "SUP-NOWHERE" in text
        assert "yield_pct" in text
        assert "penut" in text
        assert "price" in text
        # And each error names an address a human can go to.
        assert "Ingredients row 2" in text
        assert "Menu row 2" in text

    def test_a_broken_file_writes_nothing(self, db, template):
        def break_one(wb):
            wb["Menu"]["D2"] = -5

        with pytest.raises(CatalogImportError):
            import_catalog(db, _workbook(template, break_one))
        db.rollback()

        assert (
            db.execute(select(Supplier).where(Supplier.code == "SUP-EXAMPLE")).scalar_one_or_none()
            is None
        ), "a failed import must not leave the clean half behind"

    def test_a_dish_without_a_recipe_is_refused(self, db, template):
        """A menu item with no BOM is a picture of a dish, not a dish."""

        def orphan_the_dish(wb):
            ws = wb["BOM"]
            # Delete the three MNU-EXAMPLE component rows.
            for row in range(ws.max_row, 1, -1):
                if ws.cell(row=row, column=1).value == "MNU-EXAMPLE":
                    ws.delete_rows(row)

        with pytest.raises(CatalogImportError, match="no BOM rows"):
            import_catalog(db, _workbook(template, orphan_the_dish))

    def test_a_unit_mismatch_is_refused_not_converted(self, db, template):
        """Grams used as millilitres is a silent 1000x costing error waiting."""

        def wrong_unit(wb):
            wb["BOM"]["D4"] = "ml"  # ING-RICE is measured in g

        with pytest.raises(CatalogImportError, match="conversion is not supported"):
            import_catalog(db, _workbook(template, wrong_unit))

    def test_a_duplicate_sku_is_refused(self, db, template):
        def duplicate(wb):
            ws = wb["Menu"]
            ws.append(["MNU-EXAMPLE", "Same dish again", "Mains", 12, "", "grill", 2, 5])

        with pytest.raises(CatalogImportError, match="appears twice"):
            import_catalog(db, _workbook(template, duplicate))

    def test_an_uncostable_component_is_named(self, db, template):
        def phantom(wb):
            wb["BOM"]["B4"] = "ING-GHOST"

        with pytest.raises(CatalogImportError, match="ING-GHOST"):
            import_catalog(db, _workbook(template, phantom))


class TestReplaceMenu:
    def test_dishes_missing_from_the_file_are_retired(self, db, template):
        """--replace-menu makes the file the menu: seeded dishes deactivate."""
        seeded_active = db.execute(select(MenuItem).where(MenuItem.is_active)).scalars().all()
        assert seeded_active, "the seeded menu should be active before the import"

        summary = import_catalog(db, template, replace_menu=True)

        assert set(summary.deactivated) == {m.sku for m in seeded_active}
        db.expire_all()
        still_active = {
            m.sku for m in db.execute(select(MenuItem).where(MenuItem.is_active)).scalars()
        }
        assert still_active == {"MNU-EXAMPLE"}

    def test_without_the_flag_the_menus_coexist(self, db, template):
        before = {m.sku for m in db.execute(select(MenuItem).where(MenuItem.is_active)).scalars()}
        import_catalog(db, template)
        after = {m.sku for m in db.execute(select(MenuItem).where(MenuItem.is_active)).scalars()}
        assert after == before | {"MNU-EXAMPLE"}


class TestTheCli:
    def test_dry_run_reports_and_writes_nothing(self, db, template):
        from restaurant_ai.cli import app

        result = CliRunner().invoke(app, ["import-menu", str(template), "--dry-run"])
        assert result.exit_code == 0, result.output
        assert "MNU-EXAMPLE" in result.output
        assert "dry run" in result.output

    def test_a_broken_file_exits_nonzero_with_the_errors(self, db, template):
        from restaurant_ai.cli import app

        broken = _workbook(template, lambda wb: wb["Menu"].__setitem__("D2", -5))
        result = CliRunner().invoke(app, ["import-menu", str(broken)])
        assert result.exit_code == 1
        assert "price" in result.output

    def test_the_template_command_writes_an_openable_workbook(self, tmp_path):
        import openpyxl

        from restaurant_ai.cli import app

        out = tmp_path / "t.xlsx"
        result = CliRunner().invoke(app, ["menu-template", str(out)])
        assert result.exit_code == 0
        wb = openpyxl.load_workbook(out)
        assert {"ReadMe", "Suppliers", "Ingredients", "SubRecipes", "BOM", "Menu"} <= set(
            wb.sheetnames
        )


class TestAMenuBeforeItsRecipes:
    """A restaurant knows its prices long before it has costed a recipe.

    Refusing the menu until every ingredient is priced keeps the real prices out
    and leaves the demo ones in, which is worse than an incomplete catalog. The
    price of admitting them is that "uncosted" must never quietly read as "free".
    """

    @staticmethod
    def _orphan(wb):
        ws = wb["BOM"]
        for row in range(ws.max_row, 1, -1):
            if ws.cell(row=row, column=1).value == "MNU-EXAMPLE":
                ws.delete_rows(row)

    def test_refusing_is_still_the_default(self, db, template):
        """A catalog that claims to be complete should be."""
        with pytest.raises(CatalogImportError, match="no BOM rows"):
            import_catalog(db, _workbook(template, self._orphan))

    def test_the_refusal_names_the_way_through(self, db, template):
        with pytest.raises(CatalogImportError, match="allow_uncosted"):
            import_catalog(db, _workbook(template, self._orphan))

    def test_asked_for_it_loads_and_lists_them(self, db, template):
        summary = import_catalog(db, _workbook(template, self._orphan), allow_uncosted=True)
        assert "MNU-EXAMPLE" in summary.uncosted
        # It is on the menu and can be ordered.
        from restaurant_ai.db.models import MenuItem

        item = db.query(MenuItem).filter_by(sku="MNU-EXAMPLE").one()
        assert item.is_active

    def test_an_uncosted_dish_is_never_reported_as_costed(self, db, template):
        """The costing table is proof, so a dish it cannot prove stays out of it."""
        summary = import_catalog(db, _workbook(template, self._orphan), allow_uncosted=True)
        assert "MNU-EXAMPLE" not in {c["sku"] for c in summary.costings}

    def test_a_costed_dish_alongside_still_costs_out(self, db, template):
        """Admitting one uncosted dish must not weaken the proof for the rest."""
        summary = import_catalog(db, template, allow_uncosted=True)
        assert summary.costings
        assert not summary.uncosted
        assert all(c["plate_cost"] > 0 for c in summary.costings)


class TestAFileThatIsNotWhereYouAre:
    """A workbook downloaded from a chat is in Downloads, and the person is
    standing in the project folder wondering why a file they can see on screen
    "does not exist"."""

    def test_it_names_the_folder_it_searched(self, db, tmp_path, monkeypatch):
        from typer.testing import CliRunner

        from restaurant_ai.cli import app

        monkeypatch.chdir(tmp_path)
        result = CliRunner().invoke(app, ["import-menu", "menu.xlsx"])

        assert result.exit_code == 1
        # "No such file or directory" omits the one fact that resolves it.
        assert str(tmp_path) in result.output

    def test_a_spreadsheet_that_is_here_is_offered(self, db, tmp_path, monkeypatch):
        from typer.testing import CliRunner

        from restaurant_ai.cli import app

        (tmp_path / "actual-menu.xlsx").write_bytes(b"not really a workbook")
        monkeypatch.chdir(tmp_path)
        result = CliRunner().invoke(app, ["import-menu", "menu.xlsx"])

        assert "actual-menu.xlsx" in result.output

    def test_finding_it_in_downloads_hands_over_the_whole_command(self, db, tmp_path, monkeypatch):
        from typer.testing import CliRunner

        from restaurant_ai.cli import app

        downloads = tmp_path / "home" / "Downloads"
        downloads.mkdir(parents=True)
        (downloads / "menu.xlsx").write_bytes(b"not really a workbook")
        monkeypatch.setattr("pathlib.Path.home", classmethod(lambda cls: tmp_path / "home"))
        monkeypatch.chdir(tmp_path)

        result = CliRunner().invoke(app, ["import-menu", "menu.xlsx"])
        assert "Found it in your Downloads" in result.output
        # Not a hint — the command, ready to paste.
        assert "restaurant-ai import-menu" in result.output
        assert str(downloads / "menu.xlsx") in result.output


class TestStaffCanActuallyBeGivenToHenry:
    """`readiness` said "Henry needs who works here, their roles, and when they
    can work" while there was no sheet, no command and no way to say it. Advice
    for something that could not be done."""

    def _book(self, tmp_path, rows):
        import shutil

        from openpyxl import load_workbook

        path = tmp_path / "with-staff.xlsx"
        shutil.copy("menu/the-great-invention-menu.xlsx", path)
        wb = load_workbook(path)
        for row in rows:
            wb["Staff"].append(row)
        wb.save(path)
        return str(path)

    def test_a_staff_row_becomes_someone_henry_can_roster(self, db, tmp_path):
        from sqlalchemy import select

        from restaurant_ai.db.catalog_import import import_catalog
        from restaurant_ai.db.models import Staff

        path = self._book(
            tmp_path,
            [["EMP-900", "Test Person", "server", 9.5, 48, 11, "", "0,1,2", "10:00", "23:00"]],
        )
        import_catalog(db, path, allow_uncosted=True, replace_menu=True)

        person = db.execute(select(Staff).where(Staff.employee_code == "EMP-900")).scalar_one()
        assert person.role == "server"
        assert sorted(a.weekday for a in person.availability) == [0, 1, 2]
        assert str(person.availability[0].start_time) == "10:00:00"

    def test_a_role_that_is_not_a_role_is_refused_by_name(self, db, tmp_path):
        from restaurant_ai.db.catalog_import import CatalogImportError, import_catalog

        path = self._book(
            tmp_path,
            [["EMP-901", "Test", "chief cook", 9.5, 48, 11, "", "0", "10:00", "23:00"]],
        )
        with pytest.raises(CatalogImportError) as raised:
            import_catalog(db, path, allow_uncosted=True, replace_menu=True)
        assert "chief_cook" in str(raised.value)
        assert "server" in str(raised.value), "it should list what the roles actually are"

    def test_re_importing_replaces_availability_rather_than_adding_to_it(self, db, tmp_path):
        """An edited spreadsheet is the truth, not something added to what was
        there before — otherwise a corrected roster keeps the old days too."""
        from sqlalchemy import select

        from restaurant_ai.db.catalog_import import import_catalog
        from restaurant_ai.db.models import Staff

        first = self._book(
            tmp_path,
            [["EMP-902", "Test", "server", 9.5, 48, 11, "", "0,1,2,3,4", "10:00", "23:00"]],
        )
        import_catalog(db, first, allow_uncosted=True, replace_menu=True)

        second = self._book(
            tmp_path, [["EMP-902", "Test", "server", 9.5, 48, 11, "", "5,6", "10:00", "23:00"]]
        )
        import_catalog(db, second, allow_uncosted=True, replace_menu=True)

        person = db.execute(select(Staff).where(Staff.employee_code == "EMP-902")).scalar_one()
        assert sorted(a.weekday for a in person.availability) == [5, 6]

    def test_importing_the_same_sheet_twice_is_not_an_error(self, db, tmp_path):
        """The obvious thing to do after a failed import is run it again.

        `clear()` left the removals pending and SQLAlchemy inserted the new
        availability before issuing the deletes, so the second run collided with
        the unique constraint on (staff, weekday, start) — and the message named
        a UUID rather than anything an owner could act on."""
        from restaurant_ai.db.catalog_import import import_catalog

        path = self._book(
            tmp_path,
            [["EMP-903", "Test", "server", 9.5, 48, 11, "", "0,1,2", "10:00", "23:00"]],
        )
        import_catalog(db, path, allow_uncosted=True, replace_menu=True)
        import_catalog(db, path, allow_uncosted=True, replace_menu=True)
